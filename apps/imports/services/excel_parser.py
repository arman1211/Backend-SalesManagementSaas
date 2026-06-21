import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

SHEET_HINTS = {
    "customers": ["customers"],
    "payments": ["payments"],
}


def _normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _parse_decimal(value):
    if value is None or value == "":
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0.00")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _parse_int(value):
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _parse_date(value):
    if value is None or value == "":
        return None
    if hasattr(value, "date"):
        return value.date() if hasattr(value, "hour") else value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def list_workbook_sheets(uploaded_file):
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    names = workbook.sheetnames
    workbook.close()
    uploaded_file.seek(0)
    return names


def _resolve_sheet_name(sheetnames, import_type, sheet_name=None):
    if sheet_name and sheet_name in sheetnames:
        return sheet_name
    hints = SHEET_HINTS.get(import_type, [])
    lower_map = {name.lower(): name for name in sheetnames}
    for hint in hints:
        if hint in lower_map:
            return lower_map[hint]
    return sheetnames[0] if sheetnames else None


def _alias_keys(import_type):
    aliases = (
        CUSTOMER_FIELD_ALIASES if import_type == "customers" else PAYMENT_FIELD_ALIASES
    )
    keys = set()
    for field_keys in aliases.values():
        keys.update(field_keys)
    return keys


def _detect_header_row(rows, import_type, max_scan=30):
    alias_keys = _alias_keys(import_type)
    best_row = 0
    best_score = 0
    for idx, row in enumerate(rows[:max_scan]):
        headers = [_normalize_header(cell) for cell in row]
        score = sum(1 for header in headers if header in alias_keys)
        if score > best_score:
            best_score = score
            best_row = idx
    return best_row if best_score >= 2 else 0


def _read_sheet_rows(uploaded_file, import_type, sheet_name=None):
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheetnames = list(workbook.sheetnames)
    resolved_sheet = _resolve_sheet_name(sheetnames, import_type, sheet_name)
    if not resolved_sheet:
        workbook.close()
        uploaded_file.seek(0)
        return [], [], None, 1, sheetnames

    sheet = workbook[resolved_sheet]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    uploaded_file.seek(0)

    if not rows:
        return [], [], resolved_sheet, 1, sheetnames

    header_row_index = _detect_header_row(rows, import_type)
    headers = [_normalize_header(cell) for cell in rows[header_row_index]]
    data_rows = []
    excel_row_start = header_row_index + 2
    for offset, row in enumerate(rows[header_row_index + 1 :]):
        if not any(cell not in (None, "") for cell in row):
            continue
        mapped = {}
        for header, cell in zip(headers, row):
            if header:
                mapped[header] = cell
        data_rows.append({"row_number": excel_row_start + offset, "data": mapped})

    return headers, data_rows, resolved_sheet, header_row_index + 1, sheetnames


CUSTOMER_FIELD_ALIASES = {
    "customer_code": [
        "customer_id", "customer_code", "code", "id", "cust_id",
    ],
    "name": [
        "name", "customer_name", "company", "company_name", "customer",
    ],
    "trn": ["trn", "tax_registration_number", "vat_trn", "customer_trn"],
    "email": ["email", "email_address"],
    "phone": ["phone", "mobile", "telephone", "contact"],
    "address_line_1": ["address", "address_line_1", "address1", "address_1", "street"],
    "address_line_2": ["address_line_2", "address2", "address_2"],
    "city": ["city", "emirate"],
    "country": ["country"],
    "default_payment_terms_days": ["payment_terms", "payment_terms_days", "terms"],
    "notes": ["notes", "remarks"],
}


PAYMENT_FIELD_ALIASES = {
    "customer_code": ["customer_id", "customer_code", "cust_id", "code"],
    "customer_name": ["customer_name", "customer", "name", "company"],
    "invoice_number": [
        "invoice_number", "invoice_no", "inv_no", "invoice", "inv",
    ],
    "invoice_date": ["invoice_date", "inv_date", "date", "date_of_invoice"],
    "customer_lpo": ["customer_lpo", "lpo", "po_number"],
    "drn": ["drn", "delivery_reference"],
    "payment_terms_days": ["payment_terms", "payment_terms_days", "terms"],
    "grand_total": [
        "invoice_amount", "grand_total", "amount", "inv_amount", "total",
    ],
    "retention_amount": [
        "retention_amount", "retention", "retention_amt",
        "10_credit_amount", "10_amount_for_am",
    ],
    "amount_received": [
        "amount_received", "received_amount", "rev_amount", "received", "payment",
    ],
    "payment_date": ["payment_date", "received_date", "paid_date", "due_date"],
    "remarks": ["remarks", "notes"],
    "status": ["status", "payment_status"],
    "description": [
        "description",
        "work_description",
        "scope_of_work",
        "services",
        "narration",
        "item_description",
    ],
}


def parse_customer_rows(uploaded_file, sheet_name=None):
    headers, rows, sheet, header_row, sheets = _read_sheet_rows(
        uploaded_file, "customers", sheet_name
    )
    parsed = []
    errors = []
    for row in rows:
        row_number = row["row_number"]
        data = row["data"]
        record = {}
        for field, keys in CUSTOMER_FIELD_ALIASES.items():
            for key in keys:
                if key in data and data[key] not in (None, ""):
                    record[field] = data[key]
                    break
        if not record.get("name"):
            errors.append({"row": row_number, "message": "Customer name is required."})
            continue
        if record.get("customer_code") is not None:
            record["customer_code"] = str(record["customer_code"]).strip()
        if record.get("trn") is not None:
            record["trn"] = str(record["trn"]).strip()
        if record.get("default_payment_terms_days") is not None:
            terms = _parse_int(record["default_payment_terms_days"])
            if terms is None:
                errors.append({"row": row_number, "message": "Invalid payment terms."})
                continue
            record["default_payment_terms_days"] = terms
        parsed.append({"row": row_number, "record": record})

    meta = {
        "sheet": sheet,
        "header_row": header_row,
        "headers_found": [h for h in headers if h],
        "available_sheets": sheets,
    }
    return parsed, errors, meta


def parse_payment_rows(uploaded_file, sheet_name=None):
    headers, rows, sheet, header_row, sheets = _read_sheet_rows(
        uploaded_file, "payments", sheet_name
    )
    parsed = []
    errors = []
    for row in rows:
        row_number = row["row_number"]
        data = row["data"]
        record = {}
        for field, keys in PAYMENT_FIELD_ALIASES.items():
            for key in keys:
                if key in data and data[key] not in (None, ""):
                    record[field] = data[key]
                    break
        if not record.get("customer_code") and not record.get("customer_name"):
            errors.append({"row": row_number, "message": "Customer ID or name is required."})
            continue
        invoice_number = _parse_int(record.get("invoice_number"))
        if invoice_number is None or invoice_number <= 0:
            errors.append({"row": row_number, "message": "Invoice number is required."})
            continue
        grand_total = _parse_decimal(record.get("grand_total"))
        if grand_total is None or grand_total <= 0:
            errors.append({"row": row_number, "message": "Valid invoice amount is required."})
            continue
        amount_received = _parse_decimal(record.get("amount_received"))
        if amount_received is None:
            errors.append({"row": row_number, "message": "Invalid received amount."})
            continue
        retention_amount = _parse_decimal(record.get("retention_amount"))
        if retention_amount is None:
            retention_amount = (grand_total * Decimal("0.10")).quantize(Decimal("0.01"))
        invoice_date = _parse_date(record.get("invoice_date"))
        payment_date = _parse_date(record.get("payment_date")) or invoice_date
        payment_terms = _parse_int(record.get("payment_terms_days")) or 30
        parsed.append({
            "row": row_number,
            "record": {
                "customer_code": str(record["customer_code"]).strip() if record.get("customer_code") else None,
                "customer_name": str(record.get("customer_name", "")).strip(),
                "invoice_number": invoice_number,
                "invoice_date": invoice_date.isoformat() if invoice_date else None,
                "payment_date": payment_date.isoformat() if payment_date else None,
                "customer_lpo": str(record.get("customer_lpo", "")).strip(),
                "drn": str(record.get("drn", "")).strip(),
                "payment_terms_days": payment_terms,
                "grand_total": str(grand_total),
                "retention_amount": str(retention_amount),
                "amount_received": str(amount_received),
                "remarks": str(record.get("remarks", "")).strip(),
                "status": str(record.get("status", "")).strip(),
                "description": str(record.get("description", "")).strip(),
            },
        })

    meta = {
        "sheet": sheet,
        "header_row": header_row,
        "headers_found": [h for h in headers if h],
        "available_sheets": sheets,
    }
    return parsed, errors, meta
